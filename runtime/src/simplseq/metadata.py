"""Inspect and match optional participant metadata to generated sample sheets.

The public entry point remains :func:`enrich_rows_with_metadata`.  Parsing and
matching are deliberately separated so the GUI can report workbook structure,
unparseable dates, and duplicate visits before a workflow is launched.
"""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Literal, Mapping


MONTH_ALIASES = {
    name: f"{month:02d}"
    for month, names in enumerate(
        (
            ("jan", "january"), ("feb", "february"), ("mar", "march"),
            ("apr", "april"), ("may",), ("jun", "june"),
            ("jul", "july"), ("aug", "august"),
            ("sep", "sept", "september"), ("oct", "october"),
            ("nov", "november"), ("dec", "december"),
        ),
        start=1,
    )
    for name in names
}

METADATA_FIELDS = [
    "metadata_match_status", "metadata_source", "metadata_age",
    "metadata_gender", "metadata_village", "metadata_season", "metadata_pcr",
    "metadata_species", "metadata_hemoglobin", "metadata_temperature",
    "metadata_parasite_density", "metadata_status",
]

PARTICIPANT_ALIASES = ["participant_id", "participant", "patient_id", "patient", "subject", "subject_id", "asypid"]
MONTH_COLUMN_ALIASES = ["collection_month", "month", "visit_month", "asypmonths"]
DATE_ALIASES = ["collection_date", "date", "visit_date", "sample_date", "asypdates", "incidates"]
CANONICAL_COLUMNS = {
    "metadata_age": ["age", "ages"],
    "metadata_gender": ["gender", "sex"],
    "metadata_village": ["village", "villages", "site"],
    "metadata_season": ["season"],
    "metadata_pcr": [
        "pcr", "qpcr", "pcr_result", "qpcr_result", "pcr_status", "qpcr_status",
        "molecular_result", "asypcrsmear", "aspblpcr", "asypflpcr",
    ],
    "metadata_species": ["species"],
    "metadata_hemoglobin": ["hb", "hemoglobin", "asyphb"],
    "metadata_temperature": ["temperature", "temp", "asyptemp"],
    "metadata_parasite_density": ["parasite_density", "pf_density", "density", "asyppfdensities"],
    "metadata_status": ["status", "asypstatus", "asypstatusf"],
}
DateOrder = Literal["auto", "mdy", "dmy"]
METADATA_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
METADATA_NAME_HINTS = {
    "metadata": 8,
    "participant": 5,
    "patient": 5,
    "clinical": 4,
    "cohort": 3,
    "visit": 2,
    "sampleinfo": 2,
}

METADATA_COLUMN_TARGETS = (
    "participant_id",
    "collection_date",
    "month",
    "metadata_pcr",
    "metadata_season",
    "metadata_status",
    "metadata_age",
    "metadata_gender",
    "metadata_village",
    "metadata_species",
    "metadata_hemoglobin",
    "metadata_temperature",
    "metadata_parasite_density",
)
POSITIVE_DETECTION_TOKENS = {
    "+", "1", "detected", "pcrpositive", "pos", "positive", "true", "yes",
}
NEGATIVE_DETECTION_TOKENS = {
    "-", "0", "neg", "negative", "notdetected", "pcrnegative", "false", "no", "undetected",
}


@dataclass(frozen=True)
class MetadataIssue:
    severity: Literal["info", "warning", "error"]
    code: str
    message: str
    row: int | None = None


@dataclass(frozen=True)
class MetadataRecord:
    participant_id: str
    month: str
    collection_date: str
    date_source: str
    source: str
    values: dict[str, str]
    row: int = 0

    @property
    def year_month(self) -> str:
        return self.collection_date[:7] if self.collection_date else ""


@dataclass
class MetadataCatalog:
    source: str
    sheet: str
    sheets: list[str]
    header_row: int
    columns: dict[str, str]
    available_columns: list[str] = field(default_factory=list)
    value_counts: dict[str, dict[str, int]] = field(default_factory=dict)
    records: list[MetadataRecord] = field(default_factory=list)
    issues: list[MetadataIssue] = field(default_factory=list)
    _visit_index: dict[tuple[str, str], list[MetadataRecord]] = field(default_factory=dict, repr=False)
    _participant_index: dict[str, list[MetadataRecord]] = field(default_factory=dict, repr=False)

    def candidates(self, participant: str, month: str) -> list[MetadataRecord]:
        if not self._visit_index and self.records:
            for record in self.records:
                self._visit_index.setdefault((normalize_key(record.participant_id), record.month), []).append(record)
        return self._visit_index.get((normalize_key(participant), month), [])

    def participant_records(self, participant: str) -> list[MetadataRecord]:
        if not self._participant_index and self.records:
            for record in self.records:
                self._participant_index.setdefault(normalize_key(record.participant_id), []).append(record)
        return self._participant_index.get(normalize_key(participant), [])

    def summary(self) -> dict[str, Any]:
        counts = Counter(issue.severity for issue in self.issues)
        participants = {normalize_key(record.participant_id) for record in self.records}
        return {
            "source": self.source,
            "sheet": self.sheet,
            "sheets": self.sheets,
            "header_row": self.header_row,
            "columns": self.columns,
            "available_columns": self.available_columns,
            "value_counts": self.value_counts,
            "detection_values": detection_value_options(
                self.value_counts.get("metadata_pcr", {})
            ),
            "records": len(self.records),
            "participants": len(participants),
            "issues": [asdict(issue) for issue in self.issues[:50]],
            "issue_count_total": len(self.issues),
            "issue_counts": dict(counts),
        }


def normalize_key(value: object) -> str:
    raw = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", "", raw.lower())


def normalize_detection_token(value: object) -> str:
    """Normalize a detection label without collapsing literal +/- calls."""
    raw = "" if value is None else str(value).strip().lower()
    if raw in {"+", "-"}:
        return raw
    return re.sub(r"[^a-z0-9]+", "", raw)


def normalized_identifier_collisions(values: Iterable[object]) -> dict[str, list[str]]:
    """Return distinct displayed identifiers that collapse to the same key."""
    grouped: dict[str, set[str]] = defaultdict(set)
    for value in values:
        display = _display_value(value)
        key = normalize_key(display)
        if display and key:
            grouped[key].add(display)
    return {
        key: sorted(displays, key=str.casefold)
        for key, displays in grouped.items()
        if len(displays) > 1
    }


def detection_value_options(
    counts: Mapping[str, int],
    value_map: Mapping[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Return stable PCR-result values with an explicit suggested meaning."""
    configured = {
        normalize_detection_token(raw): str(state).strip().lower()
        for raw, state in (value_map or {}).items()
        if str(raw).strip()
    }
    options: list[dict[str, Any]] = []
    for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0].lower())):
        token = normalize_detection_token(value)
        state = configured.get(token, "")
        if state not in {"positive", "negative", "ignore", "review"}:
            if token in POSITIVE_DETECTION_TOKENS:
                state = "positive"
            elif token in NEGATIVE_DETECTION_TOKENS:
                state = "negative"
            else:
                state = "review"
        options.append({"value": value, "count": int(count), "state": state})
    return options


def normalize_metadata_contract(value: object) -> dict[str, Any]:
    """Normalize the user-confirmed metadata interpretation for persistence."""
    source = value if isinstance(value, Mapping) else {}
    raw_columns = source.get("columns", {})
    columns = {
        target: str(raw_columns.get(target, "") or "").strip()
        for target in METADATA_COLUMN_TARGETS
        if isinstance(raw_columns, Mapping) and target in raw_columns
    }
    raw_detection = source.get("detection_value_map", {})
    if not isinstance(raw_detection, Mapping):
        raw_detection = {}
    detection_value_map = {
        str(raw).strip(): str(state).strip().lower()
        for raw, state in raw_detection.items()
        if str(raw).strip()
        and str(state).strip().lower() in {"positive", "negative", "ignore", "review"}
    }
    raw_excluded = source.get("excluded_status_values", [])
    excluded_status_values = sorted({
        str(item).strip() for item in raw_excluded
        if str(item).strip()
    }) if isinstance(raw_excluded, (list, tuple, set)) else []
    return {
        "schema_version": 1,
        "columns": columns,
        "detection_value_map": detection_value_map,
        "excluded_status_values": excluded_status_values,
    }


def discover_metadata_file(directory: Path | str) -> Path | None:
    """Return one unambiguous metadata file colocated with FASTQ inputs."""
    root = Path(directory).expanduser()
    try:
        candidates = [
            path
            for path in root.iterdir()
            if path.is_file()
            and path.suffix.lower() in METADATA_EXTENSIONS
            and not path.name.startswith((".", "~$"))
            and normalize_key(path.stem) not in {"samples", "samplesheet", "samplemanifest"}
        ]
    except OSError:
        return None
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        return None

    scored = []
    for path in candidates:
        normalized = normalize_key(path.stem)
        score = sum(weight for hint, weight in METADATA_NAME_HINTS.items() if hint in normalized)
        scored.append((score, path.name.lower(), path))
    scored.sort(key=lambda item: (-item[0], item[1]))
    top_score = scored[0][0]
    top = [path for score, _name, path in scored if score == top_score]
    return top[0] if top_score > 0 and len(top) == 1 else None


def infer_metadata_year(
    path: Path | str,
    sheet_name: str = "",
    *,
    date_order: DateOrder = "auto",
    column_overrides: Mapping[str, str] | None = None,
) -> str:
    """Infer a fallback year only when every dated metadata record agrees."""
    catalog = inspect_metadata(
        path,
        sheet_name,
        date_order=date_order,
        column_overrides=column_overrides,
    )
    years = {record.collection_date[:4] for record in catalog.records if record.collection_date}
    return next(iter(years)) if len(years) == 1 else ""


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def first_existing_column(columns: Iterable[str], aliases: list[str]) -> str:
    lookup = {normalize_key(column): column for column in columns}
    return next((lookup[normalize_key(alias)] for alias in aliases if normalize_key(alias) in lookup), "")


def normalize_month(value: object) -> str:
    raw = _display_value(value)
    if not raw:
        return ""
    if raw.lower() in MONTH_ALIASES:
        return MONTH_ALIASES[raw.lower()]
    match = re.search(r"(?:^|\D)(0?[1-9]|1[0-2])(?:\D|$)", raw)
    return f"{int(match.group(1)):02d}" if match else ""


def parse_date(value: object, date_order: DateOrder = "auto") -> str:
    """Return ISO date, leaving ambiguous slash dates unresolved in auto mode."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    raw = _display_value(value)
    if not raw or raw.lower() in {"nan", "nat", "none", "na"}:
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-]((?:19|20)\d{2})", raw)
    if match:
        first, second, year = map(int, match.groups())
        if first > 12:
            day, month = first, second
        elif second > 12:
            month, day = first, second
        elif date_order == "mdy":
            month, day = first, second
        elif date_order == "dmy":
            day, month = first, second
        else:
            return ""
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    return ""


def month_from_date(value: object, date_order: DateOrder = "auto") -> str:
    parsed = parse_date(value, date_order)
    return parsed[5:7] if parsed else ""


def _header_score(row: list[object]) -> int:
    keys = {normalize_key(value) for value in row if _display_value(value)}
    groups = (PARTICIPANT_ALIASES, MONTH_COLUMN_ALIASES, DATE_ALIASES)
    score = sum(5 for aliases in groups if any(normalize_key(alias) in keys for alias in aliases))
    score += sum(1 for aliases in CANONICAL_COLUMNS.values() if any(normalize_key(alias) in keys for alias in aliases))
    return score


def _generic_header_score(row: list[object]) -> tuple[int, int, int]:
    """Prefer the first wide, text-like row when study labels are unfamiliar."""
    values = [value for value in row if _display_value(value)]
    if len(values) < 2:
        return (-1, -1, -1)
    text_values = sum(isinstance(value, str) for value in values)
    unique_values = len({normalize_key(value) for value in values})
    return (len(values), text_values, unique_values)


def _rows_to_records(raw_rows: list[list[object]]) -> tuple[list[dict[str, object]], int]:
    candidates = raw_rows[:20]
    if not candidates:
        return [], 0
    header_index = max(range(len(candidates)), key=lambda index: _header_score(candidates[index]))
    if _header_score(candidates[header_index]) < 5:
        header_index = max(
            range(len(candidates)),
            key=lambda index: _generic_header_score(candidates[index]),
        )
    headers = [_display_value(value) for value in raw_rows[header_index]]
    rows: list[dict[str, object]] = []
    for source_row, values in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        record = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(_display_value(value) for value in record.values()):
            record["__source_row__"] = source_row
            rows.append(record)
    return rows, header_index + 1


def _read_workbook(path: Path, sheet_name: str) -> tuple[list[dict[str, object]], int, str, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel metadata requires openpyxl. Reinstall or update the managed runtime.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = list(workbook.sheetnames)
    selected = sheet_name or ("All" if "All" in sheets else sheets[0])
    if selected not in sheets:
        raise ValueError(f"Worksheet '{selected}' was not found. Available sheets: {', '.join(sheets)}")
    worksheet = workbook[selected]
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    rows, header_row = _rows_to_records(raw_rows)
    workbook.close()
    return rows, header_row, selected, sheets


def _read_delimited(path: Path) -> tuple[list[dict[str, object]], int, str, list[str]]:
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise ValueError("Could not decode metadata file as UTF-8, UTF-16, or Windows-1252 text.")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    raw_rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    rows, header_row = _rows_to_records(raw_rows)
    return rows, header_row, "", []


def inspect_metadata(
    path: Path | str,
    sheet_name: str = "",
    *,
    date_order: DateOrder = "auto",
    column_overrides: Mapping[str, str] | None = None,
) -> MetadataCatalog:
    if date_order not in {"auto", "mdy", "dmy"}:
        raise ValueError("date_order must be auto, mdy, or dmy.")
    metadata_path = Path(path)
    if not metadata_path.is_file():
        raise ValueError(f"Metadata file does not exist: {metadata_path}")
    if metadata_path.suffix.lower() in {".xlsx", ".xlsm"}:
        rows, header_row, selected_sheet, sheets = _read_workbook(metadata_path, sheet_name)
    elif metadata_path.suffix.lower() in {".csv", ".tsv", ".tab"}:
        rows, header_row, selected_sheet, sheets = _read_delimited(metadata_path)
    else:
        raise ValueError("Metadata must be a CSV, TSV, XLSX, or XLSM file.")
    if not rows:
        return MetadataCatalog(metadata_path.name, selected_sheet, sheets, header_row, {}, issues=[
            MetadataIssue("error", "empty_table", "No metadata records were found.")
        ])

    columns = [column for column in rows[0] if column != "__source_row__"]
    participant_col = first_existing_column(columns, PARTICIPANT_ALIASES)
    month_col = first_existing_column(columns, MONTH_COLUMN_ALIASES)
    date_col = first_existing_column(columns, DATE_ALIASES)
    mapping = {"participant_id": participant_col, "month": month_col, "collection_date": date_col}
    mapping.update({name: first_existing_column(columns, aliases) for name, aliases in CANONICAL_COLUMNS.items()})
    for target, source in (column_overrides or {}).items():
        if target not in METADATA_COLUMN_TARGETS:
            continue
        requested = str(source or "").strip()
        if requested and requested not in columns:
            raise ValueError(
                f"Mapped metadata column '{requested}' for {target} was not found in "
                f"worksheet '{selected_sheet or metadata_path.name}'."
            )
        mapping[target] = requested
    participant_col = mapping.get("participant_id", "")
    month_col = mapping.get("month", "")
    date_col = mapping.get("collection_date", "")
    value_counts: dict[str, dict[str, int]] = {}
    for target in ("metadata_pcr", "metadata_status", "metadata_season"):
        source = mapping.get(target, "")
        if source:
            counts = Counter(
                _display_value(row.get(source)) for row in rows
                if _display_value(row.get(source))
            )
            value_counts[target] = dict(counts.most_common(100))
    catalog = MetadataCatalog(
        metadata_path.name,
        selected_sheet,
        sheets,
        header_row,
        mapping,
        available_columns=columns,
        value_counts=value_counts,
    )
    if not participant_col:
        catalog.issues.append(MetadataIssue("error", "participant_column_missing", "No participant or patient ID column was detected."))
        return catalog
    if not month_col and not date_col:
        catalog.issues.append(MetadataIssue("error", "visit_column_missing", "No collection date or visit month column was detected."))
        return catalog

    participant_collisions = normalized_identifier_collisions(
        row.get(participant_col) for row in rows
    )
    for displays in participant_collisions.values():
        catalog.issues.append(MetadataIssue(
            "error",
            "participant_id_collision",
            (
                "Participant identifiers normalize to the same matching key: "
                f"{', '.join(displays)}. Make these identifiers unambiguous before continuing."
            ),
        ))

    ambiguous_date_rows: list[int] = []
    for row in rows:
        source_row = int(row.get("__source_row__", 0) or 0)
        participant = _display_value(row.get(participant_col))
        if not participant:
            continue
        raw_date = row.get(date_col) if date_col else None
        parsed_date = parse_date(raw_date, date_order)
        if _display_value(raw_date) and not parsed_date:
            ambiguous_date_rows.append(source_row)
        month = month_from_date(parsed_date) or (normalize_month(row.get(month_col)) if month_col else "")
        if not month:
            catalog.issues.append(MetadataIssue("warning", "visit_month_missing", f"No usable visit month for participant {participant}.", source_row))
            continue
        values = {name: _display_value(row.get(source)) if source else "" for name, source in mapping.items() if name.startswith("metadata_")}
        catalog.records.append(MetadataRecord(participant, month, parsed_date, date_col if parsed_date else "", metadata_path.name, values, source_row))

    if ambiguous_date_rows:
        catalog.issues.append(MetadataIssue(
            "warning", "dates_unparsed",
            f"{len(ambiguous_date_rows)} date value(s) were ambiguous or invalid; visit month was used when available.",
        ))
    grouped: dict[tuple[str, str], list[MetadataRecord]] = defaultdict(list)
    for record in catalog.records:
        visit_key = record.year_month or f"month-{record.month}"
        grouped[(normalize_key(record.participant_id), visit_key)].append(record)
    for (_, visit_key), records in grouped.items():
        distinct_dates = {record.collection_date for record in records if record.collection_date}
        if len(records) > 1 and (len(distinct_dates) > 1 or len(records) != len({(record.collection_date, tuple(sorted(record.values.items()))) for record in records})):
            catalog.issues.append(MetadataIssue(
                "warning", "duplicate_visit",
                f"Participant {records[0].participant_id} has {len(records)} conflicting records for visit {visit_key}; automatic matching will avoid guessing.",
            ))
    return catalog


def read_metadata_table(path: Path, sheet_name: str = "") -> list[dict[str, str]]:
    """Compatibility helper for callers that need normalized metadata records."""
    catalog = inspect_metadata(path, sheet_name)
    return [
        {"participant_id": record.participant_id, "month": record.month, "collection_date": record.collection_date, **record.values}
        for record in catalog.records
    ]


def load_metadata_index(path: Path | str, sheet_name: str = "") -> dict[tuple[str, str], MetadataRecord]:
    """Compatibility index containing only unambiguous participant-month visits."""
    catalog = inspect_metadata(path, sheet_name)
    fatal = [issue.message for issue in catalog.issues if issue.severity == "error"]
    if fatal:
        raise ValueError(" ".join(fatal))
    grouped: dict[tuple[str, str], list[MetadataRecord]] = defaultdict(list)
    for record in catalog.records:
        grouped[(normalize_key(record.participant_id), record.month)].append(record)
    return {key: records[0] for key, records in grouped.items() if len(records) == 1}


def sample_month(sample_id: str, collection_date: str) -> str:
    date_month = month_from_date(collection_date)
    if date_month:
        return date_month
    pattern = "|".join(sorted(MONTH_ALIASES, key=len, reverse=True))
    match = re.search(rf"(^|[^A-Za-z0-9])({pattern})([^A-Za-z0-9]|$)", sample_id, re.IGNORECASE)
    return MONTH_ALIASES.get(match.group(2).lower(), "") if match else ""


def _match_record(
    catalog: MetadataCatalog,
    participant: str,
    sample_date: str,
    month: str,
    sample_id: str = "",
    sample_date_source: str = "",
) -> tuple[MetadataRecord | None, str]:
    participant_records = list(catalog.participant_records(participant))
    parsed_sample_date = parse_date(sample_date)
    source_token = str(sample_date_source or "").strip().lower()
    sample_date_is_exact = bool(parsed_sample_date) and (
        "exact_date" in source_token or source_token.startswith("metadata:")
    )
    sample_tokens = {normalize_key(token) for token in re.split(r"[^A-Za-z0-9]+", sample_id) if token}

    if sample_date_is_exact:
        exact_matches = [
            record for record in participant_records
            if record.collection_date == parsed_sample_date
        ]
        if not exact_matches:
            return (None, "date_conflict") if participant_records else (None, "missing")
        candidates = exact_matches
    else:
        candidates = list(catalog.candidates(participant, month))

    if not candidates:
        status_matches = [
            record for record in participant_records
            if normalize_key(record.values.get("metadata_status", "")) in sample_tokens
        ]
        return (status_matches[0], "matched") if len(status_matches) == 1 else (None, "missing")
    sample_year_month = parsed_sample_date[:7]
    if sample_year_month and not sample_date_is_exact:
        exact = [record for record in candidates if record.year_month == sample_year_month]
        if len(exact) == 1:
            return exact[0], "matched"
        if len(exact) > 1:
            candidates = exact
    # Some studies encode a visit/status label (for example a clinical visit)
    # in both the sample ID and metadata. Use it only when it uniquely resolves
    # otherwise ambiguous visits; never make it a required study-specific field.
    status_matches = [
        record for record in candidates
        if normalize_key(record.values.get("metadata_status", "")) in sample_tokens
    ]
    if len(status_matches) == 1:
        return status_matches[0], "matched"
    signatures = {(record.collection_date, tuple(sorted(record.values.items()))) for record in candidates}
    return (candidates[0], "matched") if len(signatures) == 1 else (None, "ambiguous")


def enrich_rows_with_metadata(
    rows: list[dict[str, str]],
    metadata_path: Path | str | None,
    *,
    metadata_sheet: str = "",
    date_order: DateOrder = "auto",
    column_overrides: Mapping[str, str] | None = None,
) -> list[dict[str, str]]:
    if not metadata_path:
        return rows
    catalog = inspect_metadata(
        metadata_path,
        metadata_sheet,
        date_order=date_order,
        column_overrides=column_overrides,
    )
    fatal = [issue.message for issue in catalog.issues if issue.severity == "error"]
    if fatal:
        raise ValueError(" ".join(fatal))
    enriched: list[dict[str, str]] = []
    for row in rows:
        next_row = dict(row)
        for field_name in METADATA_FIELDS:
            next_row.setdefault(field_name, "")
        month = sample_month(next_row.get("sample_id", ""), next_row.get("collection_date", ""))
        record, status = _match_record(
            catalog,
            next_row.get("participant_id", ""),
            next_row.get("collection_date", ""),
            month,
            next_row.get("sample_id", ""),
            next_row.get("collection_date_source", ""),
        )
        next_row["metadata_match_status"] = status
        if status == "date_conflict":
            next_row["date_note"] = (
                "Exact sample date did not match an exact metadata visit date; "
                "the sample date was kept unchanged."
            )
        if record:
            next_row["metadata_source"] = record.source
            next_row.update(record.values)
            if record.collection_date:
                next_row["collection_date"] = record.collection_date
                next_row["collection_date_source"] = f"metadata:{record.date_source}"
                next_row["inferred_year"] = "false"
                next_row["inferred_day"] = "false"
                next_row["date_note"] = "Collection date came from metadata."
        enriched.append(next_row)
    return enriched
